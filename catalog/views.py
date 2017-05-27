from django.shortcuts import render
from django.core.urlresolvers import reverse
from .models import Book, Author, BookInstance, Regal, Translator, Category
from .models import Language
from django.views import generic
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.urls import reverse_lazy
from django.http import HttpResponseRedirect
from .forms import CreateAuthorForm, PopForm, CatPopForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from .format_cell_data import get_names, get_last_first
from django.db.models import Q
from functools import reduce
import operator
import openpyxl



def index(request):
	"""
	shows basic Catalog data, welcome message and navigation panel
	"""
	num_books=Book.objects.count()
	num_instances=BookInstance.objects.count()
	num_authors=Author.objects.count()
	num_translators=Translator.objects.count()
	pl=Language(name="polski").name
	fr=Language(name="francuski").name
	books_fr=Book.objects.filter(language__name=fr).count()
	books_pl=Book.objects.filter(language__name=pl).count()
	ang=Language(name="angielski").name
	books_ang=Book.objects.filter(language__name=ang).count()
	niem=Language(name="niemiecki").name
	books_niem=Book.objects.filter(language__name=niem).count()

	#num_pl=Book.objects.
	context={
			'num_books': num_books,
			'num_authors':num_authors,
			'num_instances':num_instances,
			'books_fr':books_fr,
			'books_pl':books_pl,
			'books_ang':books_ang,
			'books_niem':books_niem,
	}

	return render(request, 'index.html', context)

class BookListView(generic.ListView):
	model=Book
	paginate_by=8

class BookDetailView(generic.DetailView):
	model=Book

class BookFullDetailView(BookDetailView):
	template_name='catalog/book_detail_full.html'


class RegalListView(generic.ListView):
	model=Regal

class RegalDetailView(generic.DetailView):
	model=Regal
	paginate_by=8


class AuthorDetailView(generic.DetailView):
	model=Author

class AuthorListView(generic.ListView):
	model=Author
	paginate_by=8

class BookCreate(LoginRequiredMixin, CreateView):
	model=Book
	fields=['author', 'title', 'published', 'publisher', 'category']
	redirect_field_name="book-create"

class BookDelete(DeleteView):
	model=Book
	success_url=reverse_lazy('index')

class BookUpdate(UpdateView):
	model=Book
	fields='__all__'

@login_required
#here is a comment for test_branch1
def author_create(request):
	if request.method=='POST':
		form=CreateAuthorForm(request.POST)

		if form.is_valid():
			f_name=form.cleaned_data['first_name']
			l_name=form.cleaned_data['last_name']
			next = request.POST.get('next', '/')
			author, created=Author.objects.get_or_create(first_name=f_name,
														last_name=l_name)

			return HttpResponseRedirect(next)

	else:
		form=CreateAuthorForm()

	return render(request, 'catalog/author_get_create.html', {'form':form})


class AuthorDelete(DeleteView):
	model=Author
	success_url=reverse_lazy('authors')


class BookSearchListView(BookListView):

	"""
	Search for book titles.
	"""
	#About the result: for term in query list (received through search form)
	#create a Q object than use operator.and_ to change iterable elements created by for
	#loop into bitwise value. Use reduce to go through all the elements)
	#see:http://stackoverflow.com/questions/44017772/how-to-use-q-objects-in-django
	paginate_by=10
	search_flag=True
	template_name='catalog/booksearch_list.html'
	def get_queryset(self):
		result=super(BookSearchListView, self).get_queryset()
		query=self.request.GET.get('q')
		if query:
			query_list=query.split()
			result=result.filter(reduce(operator.and_,(
								Q(title__icontains=q) for q in query_list))
								)

		return result
#widok do hurtowego pobierania danych
#

def catpop(request):


	if request.method=='POST':
		form=CatPopForm(request.POST)
		if form.is_valid():
			wb=openpyxl.load_workbook('data-preparation.xlsx')
			sheet=wb.get_sheet_by_name('Sheet2')
			for r in range (1,26):
				cat=sheet.cell(row=r, column=1).value
				cat_full=cat.split()
				category, created=Category.objects.get_or_create(
		                                             name=cat_full[0],
		                                             description=cat_full[1]
		                                             )

			return HttpResponseRedirect('/catalog')
	else:
		form=CatPopForm()

	return render(request, 'catalog/cat_pop_form.html', {'form':form})

def pop(request):
	"""
	Add entries to the library from xmls file using openpyxl.

	"""

	translator_obj_list=[]
	author_obj_list=[]
	if request.method=='POST':
		form=PopForm(request.POST)

		if form.is_valid():

			wb=openpyxl.load_workbook('data-preparation.xlsx')
			sheet=wb.get_sheet_by_name('Sheet1')
			for r in range(1,302):

				title=sheet.cell(row=r, column=1).value
				authors=sheet.cell(row=r, column=2).value
				published=sheet.cell(row=r, column=3).value
				publisher=sheet.cell(row=r, column=4).value
				categ=sheet.cell(row=r, column=5).value
				translators=sheet.cell(row=r, column=6).value
				isbn=sheet.cell(row=r, column=7).value
				regal=sheet.cell(row=r, column=8).value
				shelf=sheet.cell(row=r, column=9).value
				lang=sheet.cell(row=r, column=10).value

				authors_list=get_names(authors)

				for author in authors_list:
					l_name, f_name=get_last_first(author)
					author, created=Author.objects.get_or_create(
															first_name=f_name,
		                                                    last_name=l_name)
					author_obj_list.append(author)

				if translators==None:
					translator, created=Translator.objects.get_or_create(
													   first_name='---',
													   last_name='---'
														)
					translator_obj_list.append(translator)
				else:
					translators_list=get_names(translators)
					for translator in translators_list:
						l_name, f_name=get_last_first(translator)
						translator, created=Translator.objects.get_or_create(
		                                                   first_name=f_name,
		                                                   last_name=l_name
		                                                    )
						translator_obj_list.append(translator)
				reg, created=Regal.objects.get_or_create(name=regal)

				category, created=Category.objects.get_or_create(name=categ)
				cat_obj=category
				language, created=Language.objects.get_or_create(name=lang)
				lan_obj=language
				book=Book(title=title, published=published, publisher=publisher,
													isbn=isbn,
		                    						#language=language,
													#category=category,
													place=reg,
													shelf=shelf)
				book.save()

				book.language.add(lan_obj)
				book.category.add(cat_obj)

				for author in author_obj_list:
					book.author.add(author)

				for translator in translator_obj_list:
					book.translator.add(translator)

				author_obj_list=[]
				translator_obj_list=[]



			return HttpResponseRedirect('/catalog')

	else:
		form=PopForm()

	return render(request, 'catalog/pop_form.html', {'form':form})

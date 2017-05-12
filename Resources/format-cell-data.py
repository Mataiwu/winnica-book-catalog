"""
The code below gets data from a cell and puts them into lists. Its purpose is
to separate comma separated values which are stored in a single cell and thus
can not be exported to a csv file.
"""

import openpyxl

def get_names(names):
    name_list=[]
    comma_positions=[]
    #print ("string: ",names)

    for i in range(len(names)):
        #print(names[i], ": ", i)
        if names[i]==",":
            p=i
            comma_positions.append(p)
    if comma_positions==[]:
        name_list.append(names)
    else:
        y=0
        for x in comma_positions:
            name_list.append(names[y:x])
            y=x+1
        name_list.append(names[comma_positions[-1]+1:])
    return name_list



def get_first_last(name):
        if name[0]==" ":
            name=name[1:]
        if " " in name:
            for i in name:
                    
                if i==" ":
                    print("Index: ", name.index(i))
                    last_name=name[:name.index(i)]
                    first_name=name[name.index(i)+1:]
                    return last_name, first_name
                
        else:
            print("indeks_no_comma")
            last_name=name
            first_name=" "
            return last_name, first_name











#function to get data from cells in order to populate the winnica-catalgo db.
#
#def assign_values():
#    wb=openpyxl.load_workbook('data-preparation.xlsx')
#    sheet=wb.get_sheet_by_name('Sheet1')
#    for r in range(304):
#        title=sheet.cell(row=r, column=1).value
#        author=sheet.cell(row=r, column=2).value
#        published=sheet.cell(row=r, column=3).value
#        publsiher=sheet.cell(row=r, column=4).value
#        cathegory=sheet.cell(row=r, column=5).value
#        translators=sheet.cell(row=r, column=6).value
#        isbn=sheet.cell(row=r, column=7).value
#        regal=sheet.cell(row=r, column=8).value
#        shelf=sheet.cell(row=r, column=9).value
#        language=sheet.cell(row=r, column=10).value
#
def author_s_format(author_s):
    pass
#    for authors in author_s:
        #here we retreive the name from the surname
#        author, created=Author.object.get_or_create(first_name=" ", last_name=" ")
#    return author or created
#book=Book.object.create(title=title, author=author_s_format(author_s), published=published,
#                        publisher=publisher)

wb=openpyxl.load_workbook('data-preparation.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
authors= sheet.cell(row=56, column=2).value
author_s=get_names(authors)
print(author_s)
for auth in author_s:
    print("sprawdam: ", auth)
    l_name, f_name=get_first_last(auth)
    print("nazwisko %s, imię: "% l_name, f_name)

print("Author1: ", author_s[0], "Autor2: ", author_s[1])
translators=sheet.cell(row=56, column=6).value

translator_s=get_names(translators)
print("tłumacz: %s" % translator_s[0])
